"""
Unithread App — Flask backend.
REST API for expenses, revenue, receipts, calendar, chat, CRM, quotes, admin.
"""

import os
import io
import json
import uuid
import secrets
import re
from pathlib import Path
from datetime import datetime, date, timedelta
from functools import wraps

import bcrypt
from flask import (
    Flask, request, jsonify, session, render_template,
    send_from_directory, redirect, url_for, send_file
)
from flask_wtf.csrf import CSRFProtect, generate_csrf
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_socketio import SocketIO, emit, join_room, leave_room
from werkzeug.utils import secure_filename

from google_sheets import db

# ---------------------------------------------------------------------------
# App setup
# ---------------------------------------------------------------------------
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY") or secrets.token_hex(32)
app.config["UPLOAD_FOLDER"] = Path(__file__).parent / "uploads" / "receipts"
app.config["PROJECT_UPLOAD_FOLDER"] = Path(__file__).parent / "uploads" / "projects"
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16 MB
app.config["WTF_CSRF_CHECK_DEFAULT"] = False  # We check manually on mutating routes

# CSRF protection
csrf = CSRFProtect(app)

# Rate limiting
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per minute"],
    storage_uri="memory://",
)

@limiter.request_filter
def _skip_rate_limit_in_tests():
    """Disable rate limiting during automated tests."""
    return app.config.get('TESTING', False)

@app.after_request
def inject_csrf_token(response):
    """Set CSRF token cookie for JS to read."""
    if 'text/html' in response.content_type:
        response.set_cookie('csrf_token', generate_csrf(), samesite='Lax')
    return response

# SocketIO for real-time chat
socketio = SocketIO(app, cors_allowed_origins="*", manage_session=False)

BUSINESSES = ["Unithread", "Merchoteket"]

EXPENSE_CATEGORIES = [
    "Varuinköp", "Marknadsföring", "IT & Programvara", "Lokalhyra",
    "Transport & Logistik", "Design & Produktion", "Juridik & Konsulter",
    "Bank & Avgifter", "Övrigt",
]
REVENUE_CATEGORIES = [
    "Produktförsäljning", "Tjänster", "Konsultarvode", "Övrigt",
]
RECEIPT_CATEGORIES = EXPENSE_CATEGORIES
CALENDAR_TYPES = ["Möte", "Deadline", "Påminnelse", "Betalning", "Projektdeadline", "Övrigt"]
PROJECT_STATUSES = ["Aktivt", "Pausat", "Avslutat"]
TASK_STATUSES = ["Att göra", "Pågår", "Klar"]
TASK_PRIORITIES = ["Hög", "Medel", "Låg"]
VAT_RATES = [0, 6, 12, 25]

# CRM
CUSTOMER_STAGES = ["Lead", "Kontaktad", "Offert skickad", "Förhandling", "Vunnen", "Förlorad"]
CUSTOMER_SOURCES = ["Hemsida", "Referens", "LinkedIn", "Mässa", "Kall kontakt", "Övrigt"]
QUOTE_STATUSES = ["Utkast", "Skickad", "Accepterad", "Avvisad", "Fakturerad"]
INVOICE_STATUSES = ["Obetald", "Betald", "Förfallen", "Krediterad"]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _hash_password(pw):
    """Hash password with bcrypt (secure, salted)."""
    return bcrypt.hashpw(pw.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


def _verify_password(pw, hashed):
    """Verify password against hash. Supports both bcrypt and legacy SHA-256."""
    if hashed.startswith('$2b$') or hashed.startswith('$2a$'):
        return bcrypt.checkpw(pw.encode('utf-8'), hashed.encode('utf-8'))
    # Legacy SHA-256 fallback — verify then upgrade
    import hashlib
    return hashlib.sha256(pw.encode()).hexdigest() == hashed


def _upgrade_password_if_needed(users, user_dict, pw):
    """Upgrade legacy SHA-256 hash to bcrypt on successful login."""
    hashed = user_dict.get('password_hash', '')
    if not (hashed.startswith('$2b$') or hashed.startswith('$2a$')):
        user_dict['password_hash'] = _hash_password(pw)
        db.save_data('users', users)


def _validate_amount(value, field_name='belopp'):
    """Validate a monetary amount. Returns float or raises ValueError."""
    try:
        amount = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"Ogiltigt värde för {field_name}")
    if amount < 0:
        raise ValueError(f"{field_name} kan inte vara negativt")
    if amount > 999_999_999:
        raise ValueError(f"{field_name} är för stort")
    return amount


def _sanitize_string(value, max_length=500):
    """Sanitize a string input: strip, limit length."""
    if not isinstance(value, str):
        return str(value) if value is not None else ''
    return value.strip()[:max_length]


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return jsonify({"error": "Ej inloggad"}), 401
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return jsonify({"error": "Ej inloggad"}), 401
        if session.get("role") != "admin":
            return jsonify({"error": "Behörighet saknas"}), 403
        return f(*args, **kwargs)
    return decorated


def _log_activity(user, action, details=""):
    try:
        entry = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "user": user,
            "action": action,
            "details": details,
        }
        logs = db.load_data("activity_log")
        logs.append(entry)
        if len(logs) > 200:
            logs = logs[-200:]
        db.save_data("activity_log", logs)
    except Exception:
        pass


def _parse_permissions(raw):
    if not raw:
        return []
    if isinstance(raw, list):
        return raw
    if isinstance(raw, str):
        raw = raw.strip()
        if raw.startswith("["):
            try:
                return json.loads(raw)
            except Exception:
                return []
        return [p.strip() for p in raw.split(",") if p.strip()]
    return []


# ---------------------------------------------------------------------------
# Page routes
# ---------------------------------------------------------------------------

@app.route("/favicon.ico")
def favicon():
    return send_from_directory(Path(__file__).parent / "static", "favicon.ico")


@app.route("/")
def index():
    if "user" not in session:
        return render_template("login.html")
    return render_template("app.html",
                           user=session["user"],
                           role=session.get("role", "user"))


@app.route("/uploads/<path:filename>")
@login_required
def uploaded_file(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


# ---------------------------------------------------------------------------
# Auth API
# ---------------------------------------------------------------------------

@app.route("/api/login", methods=["POST"])
@limiter.limit("10 per minute")
@csrf.exempt
def api_login():
    data = request.get_json(force=True)
    username = _sanitize_string(data.get("username", ""), 50)
    password = data.get("password", "")

    if not username or not password:
        return jsonify({"ok": False, "error": "Ange användarnamn och lösenord"}), 400

    users = db.load_data("users")
    for u in users:
        if u.get("username") == username:
            if _verify_password(password, u.get("password_hash", "")):
                _upgrade_password_if_needed(users, u, password)
                session["user"] = username
                session["role"] = u.get("role", "user")
                session["permissions"] = _parse_permissions(u.get("permissions"))
                _log_activity(username, "Loggade in")
                return jsonify({"ok": True, "user": username, "role": session["role"]})
            return jsonify({"ok": False, "error": "Fel lösenord"}), 401
    return jsonify({"ok": False, "error": "Användaren finns inte"}), 401


@app.route("/api/logout", methods=["POST"])
def api_logout():
    user = session.get("user")
    if user:
        _log_activity(user, "Loggade ut")
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/me")
@login_required
def api_me():
    return jsonify({
        "user": session["user"],
        "role": session.get("role", "user"),
        "permissions": session.get("permissions", []),
    })


@app.route("/api/users-list")
@login_required
def api_users_list():
    """Return just usernames (for chat member selection etc.)"""
    users = db.load_data("users")
    return jsonify([u.get("username") for u in users if u.get("username")])


@app.route("/api/constants")
@login_required
def api_constants():
    return jsonify({
        "businesses": BUSINESSES,
        "expense_categories": EXPENSE_CATEGORIES,
        "revenue_categories": REVENUE_CATEGORIES,
        "receipt_categories": RECEIPT_CATEGORIES,
        "calendar_types": CALENDAR_TYPES,
        "vat_rates": VAT_RATES,
        "customer_stages": CUSTOMER_STAGES,
        "customer_sources": CUSTOMER_SOURCES,
        "quote_statuses": QUOTE_STATUSES,
        "invoice_statuses": INVOICE_STATUSES,
        "project_statuses": PROJECT_STATUSES,
        "task_statuses": TASK_STATUSES,
        "task_priorities": TASK_PRIORITIES,
    })


# ---------------------------------------------------------------------------
# Dashboard API
# ---------------------------------------------------------------------------

@app.route("/api/dashboard")
@login_required
def api_dashboard():
    expenses = db.load_data("expenses")
    revenue = db.load_data("revenue")
    goals = db.load_data("goals")
    receipts = db.load_data("receipts")
    logs = db.load_data("activity_log")

    now = datetime.now()
    current_month = now.strftime("%Y-%m")

    # Summaries per business
    summary = {}
    for biz in BUSINESSES:
        biz_exp = [e for e in expenses if e.get("bolag") == biz]
        biz_rev = [r for r in revenue if r.get("bolag") == biz]

        month_exp = sum(
            float(e.get("belopp", 0))
            for e in biz_exp
            if str(e.get("datum", "")).startswith(current_month)
        )
        month_rev = sum(
            float(r.get("belopp", 0))
            for r in biz_rev
            if str(r.get("datum", "")).startswith(current_month)
        )
        total_exp = sum(float(e.get("belopp", 0)) for e in biz_exp)
        total_rev = sum(float(r.get("belopp", 0)) for r in biz_rev)

        biz_goal = next((g for g in goals if g.get("bolag") == biz), {})

        summary[biz] = {
            "month_expenses": month_exp,
            "month_revenue": month_rev,
            "total_expenses": total_exp,
            "total_revenue": total_rev,
            "profit": total_rev - total_exp,
            "annual_revenue_goal": float(biz_goal.get("annual_revenue", 0)),
            "annual_profit_goal": float(biz_goal.get("annual_profit", 0)),
        }

    # Expense breakdown by category (current year)
    year = str(now.year)
    cat_totals = {}
    for e in expenses:
        if str(e.get("datum", "")).startswith(year):
            cat = e.get("kategori", "Övrigt")
            cat_totals[cat] = cat_totals.get(cat, 0) + float(e.get("belopp", 0))

    # Monthly trend (last 6 months)
    monthly = []
    for i in range(5, -1, -1):
        m = now.month - i
        y = now.year
        while m <= 0:
            m += 12
            y -= 1
        key = f"{y}-{m:02d}"
        m_exp = sum(float(e.get("belopp", 0)) for e in expenses if str(e.get("datum", "")).startswith(key))
        m_rev = sum(float(r.get("belopp", 0)) for r in revenue if str(r.get("datum", "")).startswith(key))
        monthly.append({"month": key, "expenses": m_exp, "revenue": m_rev})

    # Pending receipts
    pending_count = len([r for r in receipts if r.get("status") == "inlamnat"])

    # Budget vs actual per category (for dashboard chart)
    budget_data = db.load_data("budget")
    budget_vs_actual = {}
    for biz in BUSINESSES:
        biz_budget = next((b for b in budget_data if b.get("bolag") == biz), {})
        cats_raw = biz_budget.get("kategorier", "{}")
        if isinstance(cats_raw, str):
            try:
                cats = json.loads(cats_raw)
            except Exception:
                cats = {}
        else:
            cats = cats_raw
        biz_year_exp = [e for e in expenses if e.get("bolag") == biz and str(e.get("datum", "")).startswith(year)]
        budget_vs_actual[biz] = {}
        for cat in EXPENSE_CATEGORIES:
            cat_budget = float(cats.get(cat, 0))
            cat_spent = sum(float(e.get("belopp", 0)) for e in biz_year_exp if e.get("kategori") == cat)
            if cat_budget > 0 or cat_spent > 0:
                budget_vs_actual[biz][cat] = {"budget": cat_budget, "spent": cat_spent}

    # Revenue by category
    rev_cat_totals = {}
    for r in revenue:
        if str(r.get("datum", "")).startswith(year):
            cat = r.get("kategori", "Övrigt")
            rev_cat_totals[cat] = rev_cat_totals.get(cat, 0) + float(r.get("belopp", 0))

    return jsonify({
        "summary": summary,
        "category_breakdown": cat_totals,
        "revenue_breakdown": rev_cat_totals,
        "budget_vs_actual": budget_vs_actual,
        "monthly_trend": monthly,
        "pending_receipts": pending_count,
        "recent_activity": (logs[-10:] if logs else [])[::-1],
    })


# ---------------------------------------------------------------------------
# Expenses API
# ---------------------------------------------------------------------------

@app.route("/api/expenses", methods=["GET"])
@login_required
def get_expenses():
    data = db.load_data("expenses")
    bolag = request.args.get("bolag")
    month = request.args.get("month")  # YYYY-MM
    if bolag and bolag != "Alla":
        data = [e for e in data if e.get("bolag") == bolag]
    if month:
        data = [e for e in data if str(e.get("datum", "")).startswith(month)]
    # Sort by date desc
    data.sort(key=lambda x: x.get("datum", ""), reverse=True)
    return jsonify(data)


@app.route("/api/expenses", methods=["POST"])
@login_required
def add_expense():
    d = request.get_json(force=True)
    try:
        belopp = _validate_amount(d.get("belopp", 0))
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400

    bolag = _sanitize_string(d.get("bolag", BUSINESSES[0]), 50)
    if bolag not in BUSINESSES:
        return jsonify({"ok": False, "error": "Ogiltigt bolag"}), 400

    moms_sats = int(d.get("moms_sats", 25))
    if moms_sats not in VAT_RATES:
        return jsonify({"ok": False, "error": "Ogiltig momssats"}), 400

    expense = {
        "id": str(uuid.uuid4())[:8],
        "bolag": bolag,
        "datum": _sanitize_string(d.get("datum", date.today().isoformat()), 10),
        "kategori": _sanitize_string(d.get("kategori", "Övrigt"), 50),
        "beskrivning": _sanitize_string(d.get("beskrivning", ""), 200),
        "leverantor": _sanitize_string(d.get("leverantor", ""), 100),
        "belopp": belopp,
        "moms_sats": moms_sats,
    }
    expense["moms_belopp"] = round(
        expense["belopp"] * expense["moms_sats"] / (100 + expense["moms_sats"]), 2
    )
    db.append_row("expenses", expense)
    _log_activity(session["user"], "Lade till utgift",
                  f"{expense['beskrivning']} — {expense['belopp']} kr ({expense['bolag']})")
    return jsonify({"ok": True, "expense": expense})


@app.route("/api/expenses/<eid>", methods=["PUT"])
@login_required
def update_expense(eid):
    updates = request.get_json(force=True)
    data = db.load_data("expenses")
    for row in data:
        if str(row.get("id")) == eid:
            row.update(updates)
            if "belopp" in updates or "moms_sats" in updates:
                belopp = float(row.get("belopp", 0))
                sats = int(row.get("moms_sats", 25))
                row["moms_belopp"] = round(belopp * sats / (100 + sats), 2)
            break
    db.save_data("expenses", data)
    _log_activity(session["user"], "Uppdaterade utgift", eid)
    return jsonify({"ok": True})


@app.route("/api/expenses/<eid>", methods=["DELETE"])
@login_required
def delete_expense(eid):
    data = db.load_data("expenses")
    data = [e for e in data if str(e.get("id")) != eid]
    db.save_data("expenses", data)
    _log_activity(session["user"], "Raderade utgift", eid)
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Revenue API
# ---------------------------------------------------------------------------

@app.route("/api/revenue", methods=["GET"])
@login_required
def get_revenue():
    data = db.load_data("revenue")
    bolag = request.args.get("bolag")
    month = request.args.get("month")
    if bolag and bolag != "Alla":
        data = [r for r in data if r.get("bolag") == bolag]
    if month:
        data = [r for r in data if str(r.get("datum", "")).startswith(month)]
    data.sort(key=lambda x: x.get("datum", ""), reverse=True)
    return jsonify(data)


@app.route("/api/revenue", methods=["POST"])
@login_required
def add_revenue():
    d = request.get_json(force=True)
    try:
        belopp = _validate_amount(d.get("belopp", 0))
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400

    bolag = _sanitize_string(d.get("bolag", BUSINESSES[0]), 50)
    if bolag not in BUSINESSES:
        return jsonify({"ok": False, "error": "Ogiltigt bolag"}), 400

    rev = {
        "id": str(uuid.uuid4())[:8],
        "bolag": bolag,
        "datum": _sanitize_string(d.get("datum", date.today().isoformat()), 10),
        "kategori": _sanitize_string(d.get("kategori", "Övrigt"), 50),
        "beskrivning": _sanitize_string(d.get("beskrivning", ""), 200),
        "kund": _sanitize_string(d.get("kund", ""), 100),
        "belopp": belopp,
    }
    db.append_row("revenue", rev)
    _log_activity(session["user"], "Lade till intäkt",
                  f"{rev['beskrivning']} — {rev['belopp']} kr ({rev['bolag']})")
    return jsonify({"ok": True, "revenue": rev})


@app.route("/api/revenue/<rid>", methods=["DELETE"])
@login_required
def delete_revenue(rid):
    data = db.load_data("revenue")
    data = [r for r in data if str(r.get("id")) != rid]
    db.save_data("revenue", data)
    _log_activity(session["user"], "Raderade intäkt", rid)
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Budget API
# ---------------------------------------------------------------------------

@app.route("/api/budget", methods=["GET"])
@login_required
def get_budget():
    data = db.load_data("budget")
    result = {}
    for row in data:
        biz = row.get("bolag")
        if biz:
            result[biz] = {
                "total": float(row.get("total", 0)),
                "kategorier": json.loads(row["kategorier"]) if isinstance(row.get("kategorier"), str) and row["kategorier"].strip().startswith("{") else {},
            }
    return jsonify(result)


@app.route("/api/budget", methods=["POST"])
@login_required
def save_budget():
    d = request.get_json(force=True)
    bolag = d.get("bolag")
    budget_row = {
        "bolag": bolag,
        "total": float(d.get("total", 0)),
        "kategorier": json.dumps(d.get("kategorier", {}), ensure_ascii=False),
    }
    data = db.load_data("budget")
    found = False
    for row in data:
        if row.get("bolag") == bolag:
            row.update(budget_row)
            found = True
            break
    if not found:
        data.append(budget_row)
    db.save_data("budget", data)
    _log_activity(session["user"], "Uppdaterade budget", bolag)
    return jsonify({"ok": True})


@app.route("/api/budget/warnings")
@login_required
def get_budget_warnings():
    """Return categories that exceed 80% or 100% of budget."""
    budget_data = db.load_data("budget")
    expenses = db.load_data("expenses")
    year = str(datetime.now().year)

    warnings = []
    for biz in BUSINESSES:
        biz_budget = next((b for b in budget_data if b.get("bolag") == biz), {})
        cats_raw = biz_budget.get("kategorier", "{}")
        if isinstance(cats_raw, str):
            try:
                cats = json.loads(cats_raw)
            except Exception:
                cats = {}
        else:
            cats = cats_raw

        total_budget = float(biz_budget.get("total", 0))
        biz_year_exp = [e for e in expenses if e.get("bolag") == biz and str(e.get("datum", "")).startswith(year)]
        total_spent = sum(float(e.get("belopp", 0)) for e in biz_year_exp)

        # Total budget warning
        if total_budget > 0:
            pct = (total_spent / total_budget) * 100
            if pct >= 100:
                warnings.append({"bolag": biz, "kategori": "Total", "budget": total_budget, "spent": total_spent, "pct": round(pct, 1), "level": "danger"})
            elif pct >= 80:
                warnings.append({"bolag": biz, "kategori": "Total", "budget": total_budget, "spent": total_spent, "pct": round(pct, 1), "level": "warning"})

        # Per category
        for cat in EXPENSE_CATEGORIES:
            cat_budget = float(cats.get(cat, 0))
            if cat_budget <= 0:
                continue
            cat_spent = sum(float(e.get("belopp", 0)) for e in biz_year_exp if e.get("kategori") == cat)
            pct = (cat_spent / cat_budget) * 100
            if pct >= 100:
                warnings.append({"bolag": biz, "kategori": cat, "budget": cat_budget, "spent": cat_spent, "pct": round(pct, 1), "level": "danger"})
            elif pct >= 80:
                warnings.append({"bolag": biz, "kategori": cat, "budget": cat_budget, "spent": cat_spent, "pct": round(pct, 1), "level": "warning"})

    warnings.sort(key=lambda w: w["pct"], reverse=True)
    return jsonify(warnings)


# ---------------------------------------------------------------------------
# Goals API
# ---------------------------------------------------------------------------

@app.route("/api/goals", methods=["GET"])
@login_required
def get_goals():
    data = db.load_data("goals")
    result = {}
    for row in data:
        biz = row.get("bolag")
        if biz:
            result[biz] = {
                "annual_revenue": float(row.get("annual_revenue", 0)),
                "annual_profit": float(row.get("annual_profit", 0)),
            }
    return jsonify(result)


@app.route("/api/goals", methods=["POST"])
@login_required
def save_goals():
    d = request.get_json(force=True)
    bolag = d.get("bolag")
    goal = {
        "bolag": bolag,
        "annual_revenue": float(d.get("annual_revenue", 0)),
        "annual_profit": float(d.get("annual_profit", 0)),
    }
    data = db.load_data("goals")
    found = False
    for row in data:
        if row.get("bolag") == bolag:
            row.update(goal)
            found = True
            break
    if not found:
        data.append(goal)
    db.save_data("goals", data)
    _log_activity(session["user"], "Uppdaterade mål", bolag)
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Receipts API
# ---------------------------------------------------------------------------

@app.route("/api/receipts", methods=["GET"])
@login_required
def get_receipts():
    data = db.load_data("receipts")
    status = request.args.get("status")
    bolag = request.args.get("bolag")
    month = request.args.get("month")
    if status:
        data = [r for r in data if r.get("status") == status]
    if bolag and bolag != "Alla":
        data = [r for r in data if r.get("bolag") == bolag]
    if month:
        data = [r for r in data if str(r.get("datum", "")).startswith(month)]
    data.sort(key=lambda x: x.get("created", ""), reverse=True)
    return jsonify(data)


@app.route("/api/receipts", methods=["POST"])
@login_required
def add_receipt():
    bolag = request.form.get("bolag", BUSINESSES[0])
    beskrivning = request.form.get("beskrivning", "")
    belopp = float(request.form.get("belopp", 0))
    kategori = request.form.get("kategori", "Övrigt")
    datum = request.form.get("datum", date.today().isoformat())

    # Handle file uploads
    files = request.files.getlist("files")
    file_paths = []
    for f in files:
        if f and f.filename:
            fname = f"{uuid.uuid4().hex[:8]}_{secure_filename(f.filename)}"
            save_path = app.config["UPLOAD_FOLDER"] / fname
            save_path.parent.mkdir(parents=True, exist_ok=True)
            f.save(str(save_path))
            file_paths.append(fname)

    receipt = {
        "id": str(uuid.uuid4())[:8],
        "user": session["user"],
        "bolag": bolag,
        "datum": datum,
        "beskrivning": beskrivning,
        "belopp": belopp,
        "kategori": kategori,
        "status": "inlamnat",
        "files": json.dumps(file_paths, ensure_ascii=False),
        "created": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    db.append_row("receipts", receipt)
    _log_activity(session["user"], "Laddade upp kvitto",
                  f"{beskrivning} — {belopp} kr")
    return jsonify({"ok": True, "receipt": receipt})


@app.route("/api/receipts/<rid>/status", methods=["PUT"])
@login_required
def update_receipt_status(rid):
    d = request.get_json(force=True)
    new_status = d.get("status")  # godkannt / avvisat
    data = db.load_data("receipts")
    for row in data:
        if str(row.get("id")) == rid:
            row["status"] = new_status
            break
    db.save_data("receipts", data)
    _log_activity(session["user"], f"Kvitto {new_status}", rid)
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Calendar API
# ---------------------------------------------------------------------------

@app.route("/api/calendar/events", methods=["GET"])
@login_required
def get_events():
    data = db.load_data("calendar_events")
    year = request.args.get("year")
    month = request.args.get("month")
    if year and month:
        prefix = f"{year}-{int(month):02d}"
        data = [e for e in data if str(e.get("datum", "")).startswith(prefix)]
    data.sort(key=lambda x: (x.get("datum", ""), x.get("time", "")))
    return jsonify(data)


@app.route("/api/calendar/events", methods=["POST"])
@login_required
def add_event():
    d = request.get_json(force=True)
    event = {
        "id": f"evt_{uuid.uuid4().hex[:8]}",
        "title": d.get("title", ""),
        "datum": d.get("datum", date.today().isoformat()),
        "time": d.get("time", ""),
        "type": d.get("type", "Övrigt"),
        "business": d.get("business", "Alla"),
        "beskrivning": d.get("beskrivning", ""),
        "created_by": session["user"],
        "created": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    db.append_row("calendar_events", event)
    _log_activity(session["user"], "Skapade händelse", event["title"])
    return jsonify({"ok": True, "event": event})


@app.route("/api/calendar/events/<eid>", methods=["DELETE"])
@login_required
def delete_event(eid):
    data = db.load_data("calendar_events")
    data = [e for e in data if str(e.get("id")) != eid]
    db.save_data("calendar_events", data)
    _log_activity(session["user"], "Raderade händelse", eid)
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Todos API
# ---------------------------------------------------------------------------

@app.route("/api/todos", methods=["GET"])
@login_required
def get_todos():
    data = db.load_data("todos")
    data.sort(key=lambda x: (
        str(x.get("done", "False")).lower() == "true",
        {"Hög": 0, "Medel": 1, "Låg": 2}.get(x.get("priority", "Medel"), 1),
    ))
    return jsonify(data)


@app.route("/api/todos", methods=["POST"])
@login_required
def add_todo():
    d = request.get_json(force=True)
    todo = {
        "id": f"todo_{uuid.uuid4().hex[:8]}",
        "task": d.get("task", ""),
        "priority": d.get("priority", "Medel"),
        "deadline": d.get("deadline", ""),
        "done": False,
        "created_by": session["user"],
        "created": datetime.now().strftime("%Y-%m-%d"),
    }
    db.append_row("todos", todo)
    return jsonify({"ok": True, "todo": todo})


@app.route("/api/todos/<tid>", methods=["PUT"])
@login_required
def update_todo(tid):
    d = request.get_json(force=True)
    data = db.load_data("todos")
    for row in data:
        if str(row.get("id")) == tid:
            row.update(d)
            break
    db.save_data("todos", data)
    return jsonify({"ok": True})


@app.route("/api/todos/<tid>", methods=["DELETE"])
@login_required
def delete_todo(tid):
    data = db.load_data("todos")
    data = [t for t in data if str(t.get("id")) != tid]
    db.save_data("todos", data)
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Chat API
# ---------------------------------------------------------------------------

@app.route("/api/chat/groups", methods=["GET"])
@login_required
def get_chat_groups():
    groups = db.load_data("chat_groups")
    user = session["user"]
    result = []
    for g in groups:
        members = g.get("members", "[]")
        if isinstance(members, str):
            try:
                members = json.loads(members)
            except Exception:
                members = []
        if user in members or session.get("role") == "admin":
            g["members"] = members
            result.append(g)
    return jsonify(result)


@app.route("/api/chat/groups", methods=["POST"])
@login_required
def create_chat_group():
    d = request.get_json(force=True)
    group = {
        "id": str(uuid.uuid4())[:8],
        "name": d.get("name", "Ny grupp"),
        "created_by": session["user"],
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "members": json.dumps(d.get("members", [session["user"]]), ensure_ascii=False),
        "archived": False,
    }
    db.append_row("chat_groups", group)
    _log_activity(session["user"], "Skapade chattgrupp", group["name"])
    return jsonify({"ok": True, "group": group})


@app.route("/api/chat/groups/<gid>", methods=["DELETE"])
@login_required
def delete_chat_group(gid):
    """Delete a chat group. Only admin or the creator may delete."""
    groups = db.load_data("chat_groups")
    user = session["user"]
    role = session.get("role")
    target = None
    for g in groups:
        if str(g.get("id")) == gid:
            target = g
            break
    if not target:
        return jsonify({"ok": False, "error": "Grupp hittades inte"}), 404
    if role != "admin" and target.get("created_by") != user:
        return jsonify({"ok": False, "error": "Ingen behörighet"}), 403
    groups = [g for g in groups if str(g.get("id")) != gid]
    db.save_data("chat_groups", groups)
    # Also delete messages belonging to this group
    messages = db.load_data("chat_messages")
    messages = [m for m in messages if str(m.get("group_id")) != gid]
    db.save_data("chat_messages", messages)
    _log_activity(user, "Raderade chattgrupp", target.get("name", gid))
    return jsonify({"ok": True})


@app.route("/api/chat/groups/<gid>/messages", methods=["GET"])
@login_required
def get_messages(gid):
    messages = db.load_data("chat_messages")
    msgs = [m for m in messages if str(m.get("group_id")) == gid]
    msgs.sort(key=lambda x: x.get("timestamp", ""))
    return jsonify(msgs)


@app.route("/api/chat/groups/<gid>/messages", methods=["POST"])
@login_required
def send_message(gid):
    d = request.get_json(force=True)
    msg = {
        "id": str(uuid.uuid4())[:8],
        "group_id": gid,
        "sender": session["user"],
        "content": d.get("content", ""),
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    db.append_row("chat_messages", msg)
    # Broadcast via WebSocket to all in room
    socketio.emit("new_message", msg, room=f"chat_{gid}")
    return jsonify({"ok": True, "message": msg})


    # Broadcast via WebSocket to all in room
    socketio.emit("new_message", msg, room=f"chat_{gid}")
    return jsonify({"ok": True, "message": msg})


# ---------------------------------------------------------------------------
# Projects API
# ---------------------------------------------------------------------------

@app.route("/api/projects", methods=["GET"])
@login_required
def get_projects():
    """Return projects the current user is a member of (or all for admin)."""
    projects = db.load_data("projects")
    user = session["user"]
    role = session.get("role")
    result = []
    for p in projects:
        members = p.get("members", "[]")
        if isinstance(members, str):
            try:
                members = json.loads(members)
            except Exception:
                members = []
        if user in members or role == "admin":
            p["members"] = members
            result.append(p)
    result.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    return jsonify(result)


@app.route("/api/projects", methods=["POST"])
@login_required
def create_project():
    d = request.get_json(force=True)
    members = d.get("members", [session["user"]])
    if session["user"] not in members:
        members.insert(0, session["user"])
    project = {
        "id": f"proj_{uuid.uuid4().hex[:8]}",
        "name": _sanitize_string(d.get("name", "Nytt projekt"), 100),
        "description": d.get("description", ""),
        "bolag": d.get("bolag", BUSINESSES[0]),
        "members": json.dumps(members, ensure_ascii=False),
        "created_by": session["user"],
        "status": "Aktivt",
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    db.append_row("projects", project)
    _log_activity(session["user"], "Skapade projekt", project["name"])
    return jsonify({"ok": True, "project": project})


@app.route("/api/projects/<pid>", methods=["PUT"])
@login_required
def update_project(pid):
    d = request.get_json(force=True)
    data = db.load_data("projects")
    for row in data:
        if str(row.get("id")) == pid:
            if "name" in d:
                row["name"] = _sanitize_string(d["name"], 100)
            if "description" in d:
                row["description"] = d["description"]
            if "status" in d and d["status"] in PROJECT_STATUSES:
                row["status"] = d["status"]
            if "members" in d:
                row["members"] = json.dumps(d["members"], ensure_ascii=False)
            break
    db.save_data("projects", data)
    return jsonify({"ok": True})


@app.route("/api/projects/<pid>", methods=["DELETE"])
@login_required
def delete_project(pid):
    """Delete a project. Only admin or creator may delete."""
    projects = db.load_data("projects")
    user = session["user"]
    role = session.get("role")
    target = None
    for p in projects:
        if str(p.get("id")) == pid:
            target = p
            break
    if not target:
        return jsonify({"ok": False, "error": "Projekt hittades inte"}), 404
    if role != "admin" and target.get("created_by") != user:
        return jsonify({"ok": False, "error": "Ingen behörighet"}), 403
    projects = [p for p in projects if str(p.get("id")) != pid]
    db.save_data("projects", projects)
    # Delete tasks belonging to this project
    tasks = db.load_data("project_tasks")
    tasks = [t for t in tasks if str(t.get("project_id")) != pid]
    db.save_data("project_tasks", tasks)
    # Delete files belonging to this project
    files = db.load_data("project_files")
    proj_files = [f for f in files if str(f.get("project_id")) == pid]
    for pf in proj_files:
        fpath = app.config["PROJECT_UPLOAD_FOLDER"] / pf.get("filename", "")
        if fpath.exists():
            fpath.unlink()
    files = [f for f in files if str(f.get("project_id")) != pid]
    db.save_data("project_files", files)
    # Delete related calendar events
    events = db.load_data("calendar_events")
    events = [e for e in events if e.get("project_id") != pid]
    db.save_data("calendar_events", events)
    _log_activity(user, "Raderade projekt", target.get("name", pid))
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Project Tasks API
# ---------------------------------------------------------------------------

@app.route("/api/projects/<pid>/tasks", methods=["GET"])
@login_required
def get_project_tasks(pid):
    tasks = db.load_data("project_tasks")
    tasks = [t for t in tasks if str(t.get("project_id")) == pid]
    tasks.sort(key=lambda x: (
        {"Hög": 0, "Medel": 1, "Låg": 2}.get(x.get("priority", "Medel"), 1),
        x.get("deadline", "9999"),
    ))
    return jsonify(tasks)


@app.route("/api/projects/<pid>/tasks", methods=["POST"])
@login_required
def create_project_task(pid):
    d = request.get_json(force=True)
    task = {
        "id": f"ptask_{uuid.uuid4().hex[:8]}",
        "project_id": pid,
        "title": _sanitize_string(d.get("title", ""), 200),
        "description": d.get("description", ""),
        "assigned_to": d.get("assigned_to", ""),
        "created_by": session["user"],
        "deadline": d.get("deadline", ""),
        "status": "Att göra",
        "priority": d.get("priority", "Medel") if d.get("priority") in TASK_PRIORITIES else "Medel",
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    db.append_row("project_tasks", task)
    _log_activity(session["user"], "Skapade uppgift", task["title"])
    # Create calendar event if task has a deadline
    if task["deadline"]:
        _sync_task_to_calendar(task, pid)
    return jsonify({"ok": True, "task": task})


@app.route("/api/projects/<pid>/tasks/<tid>", methods=["PUT"])
@login_required
def update_project_task(pid, tid):
    d = request.get_json(force=True)
    data = db.load_data("project_tasks")
    for row in data:
        if str(row.get("id")) == tid and str(row.get("project_id")) == pid:
            if "title" in d:
                row["title"] = _sanitize_string(d["title"], 200)
            if "description" in d:
                row["description"] = d["description"]
            if "assigned_to" in d:
                row["assigned_to"] = d["assigned_to"]
            if "deadline" in d:
                row["deadline"] = d["deadline"]
            if "status" in d and d["status"] in TASK_STATUSES:
                row["status"] = d["status"]
            if "priority" in d and d["priority"] in TASK_PRIORITIES:
                row["priority"] = d["priority"]
            # Sync calendar event
            if row.get("deadline"):
                _sync_task_to_calendar(row, pid)
            else:
                # Remove calendar event if deadline was cleared
                _remove_task_from_calendar(tid)
            break
    db.save_data("project_tasks", data)
    return jsonify({"ok": True})


@app.route("/api/projects/<pid>/tasks/<tid>", methods=["DELETE"])
@login_required
def delete_project_task(pid, tid):
    data = db.load_data("project_tasks")
    data = [t for t in data if not (str(t.get("id")) == tid and str(t.get("project_id")) == pid)]
    db.save_data("project_tasks", data)
    _remove_task_from_calendar(tid)
    return jsonify({"ok": True})


def _sync_task_to_calendar(task, project_id):
    """Create or update a calendar event for a project task deadline."""
    projects = db.load_data("projects")
    proj_name = ""
    for p in projects:
        if str(p.get("id")) == project_id:
            proj_name = p.get("name", "")
            break
    events = db.load_data("calendar_events")
    evt_id = f"ptask_cal_{task['id']}"
    title = f"📋 {task['title']}"
    if task.get("assigned_to"):
        title += f" → {task['assigned_to']}"
    found = False
    for e in events:
        if str(e.get("id")) == evt_id:
            e["title"] = title
            e["datum"] = task["deadline"]
            e["beskrivning"] = f"Projekt: {proj_name}"
            found = True
            break
    if not found:
        events.append({
            "id": evt_id,
            "title": title,
            "datum": task["deadline"],
            "time": "",
            "type": "Projektdeadline",
            "business": "Alla",
            "beskrivning": f"Projekt: {proj_name}",
            "project_id": project_id,
            "task_id": task["id"],
            "created_by": task.get("created_by", ""),
            "created": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })
    db.save_data("calendar_events", events)


def _remove_task_from_calendar(task_id):
    """Remove the calendar event linked to a project task."""
    events = db.load_data("calendar_events")
    evt_id = f"ptask_cal_{task_id}"
    events = [e for e in events if str(e.get("id")) != evt_id]
    db.save_data("calendar_events", events)


# ---------------------------------------------------------------------------
# Project Files API
# ---------------------------------------------------------------------------

@app.route("/api/projects/<pid>/files", methods=["GET"])
@login_required
def get_project_files(pid):
    files = db.load_data("project_files")
    files = [f for f in files if str(f.get("project_id")) == pid]
    files.sort(key=lambda x: x.get("uploaded_at", ""), reverse=True)
    return jsonify(files)


@app.route("/api/projects/<pid>/files", methods=["POST"])
@login_required
def upload_project_file(pid):
    """Upload a file to a project."""
    files = request.files.getlist("files")
    if not files:
        return jsonify({"ok": False, "error": "Ingen fil vald"}), 400
    uploaded = []
    for f in files:
        if f and f.filename:
            safe_name = secure_filename(f.filename)
            fname = f"{uuid.uuid4().hex[:8]}_{safe_name}"
            save_path = app.config["PROJECT_UPLOAD_FOLDER"] / fname
            save_path.parent.mkdir(parents=True, exist_ok=True)
            f.save(str(save_path))
            file_rec = {
                "id": f"pfile_{uuid.uuid4().hex[:8]}",
                "project_id": pid,
                "filename": fname,
                "original_name": f.filename,
                "uploaded_by": session["user"],
                "uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            db.append_row("project_files", file_rec)
            uploaded.append(file_rec)
    _log_activity(session["user"], "Laddade upp projektfil",
                  f"{len(uploaded)} filer till projekt {pid}")
    return jsonify({"ok": True, "files": uploaded})


@app.route("/api/projects/<pid>/files/<fid>", methods=["DELETE"])
@login_required
def delete_project_file(pid, fid):
    files = db.load_data("project_files")
    target = None
    for f in files:
        if str(f.get("id")) == fid and str(f.get("project_id")) == pid:
            target = f
            break
    if not target:
        return jsonify({"ok": False, "error": "Fil hittades inte"}), 404
    # Delete physical file
    fpath = app.config["PROJECT_UPLOAD_FOLDER"] / target.get("filename", "")
    if fpath.exists():
        fpath.unlink()
    files = [f for f in files if str(f.get("id")) != fid]
    db.save_data("project_files", files)
    return jsonify({"ok": True})


@app.route("/api/projects/files/<filename>")
@login_required
def serve_project_file(filename):
    """Serve an uploaded project file."""
    folder = app.config["PROJECT_UPLOAD_FOLDER"]
    return send_from_directory(str(folder), filename)


# ---------------------------------------------------------------------------
# WebSocket events (real-time chat)
# ---------------------------------------------------------------------------

@socketio.on("join_chat")
def handle_join_chat(data):
    """User joins a chat room for real-time updates."""
    gid = data.get("group_id")
    if gid and "user" in session:
        join_room(f"chat_{gid}")
        emit("user_joined", {"user": session["user"], "group_id": gid}, room=f"chat_{gid}")

@socketio.on("leave_chat")
def handle_leave_chat(data):
    gid = data.get("group_id")
    if gid:
        leave_room(f"chat_{gid}")

@socketio.on("typing")
def handle_typing(data):
    """Broadcast typing indicator to room."""
    gid = data.get("group_id")
    if gid and "user" in session:
        emit("user_typing", {"user": session["user"], "group_id": gid}, room=f"chat_{gid}", include_self=False)


# ---------------------------------------------------------------------------
# Admin API
# ---------------------------------------------------------------------------

@app.route("/api/admin/users", methods=["GET"])
@admin_required
def admin_get_users():
    users = db.load_data("users")
    # Don't send password hashes to client
    safe = []
    for u in users:
        safe.append({
            "username": u.get("username"),
            "role": u.get("role"),
            "permissions": _parse_permissions(u.get("permissions")),
        })
    return jsonify(safe)


@app.route("/api/admin/users", methods=["POST"])
@admin_required
def admin_create_user():
    d = request.get_json(force=True)
    username = _sanitize_string(d.get("username", ""), 50)
    password = d.get("password", "")
    role = d.get("role", "user")
    permissions = d.get("permissions", [])

    if not username:
        return jsonify({"ok": False, "error": "Användarnamn krävs"}), 400
    if not re.match(r'^[a-zA-ZåäöÅÄÖ0-9_ -]{2,50}$', username):
        return jsonify({"ok": False, "error": "Ogiltigt användarnamn"}), 400
    if not password or len(password) < 6:
        return jsonify({"ok": False, "error": "Lösenord måste vara minst 6 tecken"}), 400
    if role not in ("admin", "user"):
        return jsonify({"ok": False, "error": "Ogiltig roll"}), 400

    users = db.load_data("users")
    for u in users:
        if u.get("username") == username:
            return jsonify({"ok": False, "error": "Användaren finns redan"}), 409

    new_user = {
        "username": username,
        "password_hash": _hash_password(password),
        "role": role,
        "permissions": json.dumps(permissions, ensure_ascii=False),
    }
    users.append(new_user)
    db.save_data("users", users)
    _log_activity(session["user"], "Skapade användare", username)
    return jsonify({"ok": True})


@app.route("/api/admin/users/<username>", methods=["PUT"])
@admin_required
def admin_update_user(username):
    d = request.get_json(force=True)
    users = db.load_data("users")
    for u in users:
        if u.get("username") == username:
            if "role" in d:
                u["role"] = d["role"]
            if "permissions" in d:
                u["permissions"] = json.dumps(d["permissions"], ensure_ascii=False)
            if "password" in d and d["password"]:
                if len(d["password"]) < 6:
                    return jsonify({"ok": False, "error": "Lösenord måste vara minst 6 tecken"}), 400
                u["password_hash"] = _hash_password(d["password"])
            break
    db.save_data("users", users)
    _log_activity(session["user"], "Uppdaterade användare", username)
    return jsonify({"ok": True})


@app.route("/api/admin/users/<username>", methods=["DELETE"])
@admin_required
def admin_delete_user(username):
    if username in ("Viktor", "admin"):
        return jsonify({"ok": False, "error": "Kan inte ta bort denna användare"}), 403
    users = db.load_data("users")
    users = [u for u in users if u.get("username") != username]
    db.save_data("users", users)
    _log_activity(session["user"], "Raderade användare", username)
    return jsonify({"ok": True})


@app.route("/api/admin/activity-log")
@login_required
def get_activity_log():
    logs = db.load_data("activity_log")
    logs.reverse()
    return jsonify(logs[:100])


# ---------------------------------------------------------------------------
# CRM — Customers API
# ---------------------------------------------------------------------------

@app.route("/api/customers", methods=["GET"])
@login_required
def get_customers():
    data = db.load_data("customers")
    stage = request.args.get("stage")
    bolag = request.args.get("bolag")
    search = request.args.get("q", "").lower()
    if stage and stage != "Alla":
        data = [c for c in data if c.get("stage") == stage]
    if bolag and bolag != "Alla":
        data = [c for c in data if c.get("bolag") == bolag]
    if search:
        data = [c for c in data if search in (c.get("name", "") + c.get("email", "") + c.get("company", "")).lower()]
    data.sort(key=lambda x: x.get("updated", x.get("created", "")), reverse=True)
    return jsonify(data)


@app.route("/api/customers", methods=["POST"])
@login_required
def add_customer():
    d = request.get_json(force=True)
    customer = {
        "id": f"cust_{uuid.uuid4().hex[:8]}",
        "name": d.get("name", ""),
        "company": d.get("company", ""),
        "email": d.get("email", ""),
        "phone": d.get("phone", ""),
        "bolag": d.get("bolag", BUSINESSES[0]),
        "stage": d.get("stage", "Lead"),
        "source": d.get("source", "Övrigt"),
        "value": float(d.get("value", 0)),
        "notes": d.get("notes", ""),
        "assigned_to": d.get("assigned_to", session["user"]),
        "created": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    db.append_row("customers", customer)
    _log_activity(session["user"], "Lade till kund", customer["name"])
    return jsonify({"ok": True, "customer": customer})


@app.route("/api/customers/<cid>", methods=["PUT"])
@login_required
def update_customer(cid):
    updates = request.get_json(force=True)
    data = db.load_data("customers")
    for row in data:
        if str(row.get("id")) == cid:
            row.update(updates)
            row["updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            break
    db.save_data("customers", data)
    _log_activity(session["user"], "Uppdaterade kund", cid)
    return jsonify({"ok": True})


@app.route("/api/customers/<cid>", methods=["DELETE"])
@login_required
def delete_customer(cid):
    data = db.load_data("customers")
    data = [c for c in data if str(c.get("id")) != cid]
    db.save_data("customers", data)
    _log_activity(session["user"], "Raderade kund", cid)
    return jsonify({"ok": True})


@app.route("/api/customers/<cid>/notes", methods=["POST"])
@login_required
def add_customer_note(cid):
    d = request.get_json(force=True)
    note = {
        "id": f"note_{uuid.uuid4().hex[:6]}",
        "customer_id": cid,
        "text": d.get("text", ""),
        "author": session["user"],
        "created": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    db.append_row("customer_notes", note)
    return jsonify({"ok": True, "note": note})


@app.route("/api/customers/<cid>/notes", methods=["GET"])
@login_required
def get_customer_notes(cid):
    data = db.load_data("customer_notes")
    notes = [n for n in data if str(n.get("customer_id")) == cid]
    notes.sort(key=lambda x: x.get("created", ""), reverse=True)
    return jsonify(notes)


# ---------------------------------------------------------------------------
# CRM — Pipeline API (customers grouped by stage)
# ---------------------------------------------------------------------------

@app.route("/api/pipeline", methods=["GET"])
@login_required
def get_pipeline():
    customers = db.load_data("customers")
    bolag = request.args.get("bolag")
    if bolag and bolag != "Alla":
        customers = [c for c in customers if c.get("bolag") == bolag]

    pipeline = {}
    for stage in CUSTOMER_STAGES:
        stage_custs = [c for c in customers if c.get("stage") == stage]
        total_value = sum(float(c.get("value", 0)) for c in stage_custs)
        pipeline[stage] = {"customers": stage_custs, "count": len(stage_custs), "total_value": total_value}
    return jsonify(pipeline)


@app.route("/api/customers/<cid>/stage", methods=["PUT"])
@login_required
def update_customer_stage(cid):
    d = request.get_json(force=True)
    new_stage = d.get("stage")
    data = db.load_data("customers")
    name = ""
    for row in data:
        if str(row.get("id")) == cid:
            row["stage"] = new_stage
            row["updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            name = row.get("name", cid)
            break
    db.save_data("customers", data)
    _log_activity(session["user"], f"Flyttade kund till {new_stage}", name)
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Quotes / Offerter API
# ---------------------------------------------------------------------------

@app.route("/api/quotes", methods=["GET"])
@login_required
def get_quotes():
    data = db.load_data("quotes")
    status = request.args.get("status")
    bolag = request.args.get("bolag")
    if status and status != "Alla":
        data = [q for q in data if q.get("status") == status]
    if bolag and bolag != "Alla":
        data = [q for q in data if q.get("bolag") == bolag]
    data.sort(key=lambda x: x.get("created", ""), reverse=True)
    return jsonify(data)


@app.route("/api/quotes", methods=["POST"])
@login_required
def create_quote():
    d = request.get_json(force=True)
    items = d.get("items", [])
    subtotal = sum(float(it.get("total", 0)) for it in items)
    moms_total = sum(
        float(it.get("total", 0)) * int(it.get("moms", 25)) / (100 + int(it.get("moms", 25)))
        for it in items
    )

    quote = {
        "id": f"Q-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}",
        "customer_id": d.get("customer_id", ""),
        "customer_name": d.get("customer_name", ""),
        "bolag": d.get("bolag", BUSINESSES[0]),
        "title": d.get("title", "Offert"),
        "description": d.get("description", ""),
        "items": json.dumps(items, ensure_ascii=False),
        "subtotal": round(subtotal, 2),
        "moms_total": round(moms_total, 2),
        "total": round(subtotal, 2),
        "valid_until": d.get("valid_until", ""),
        "status": "Utkast",
        "created_by": session["user"],
        "created": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    db.append_row("quotes", quote)
    _log_activity(session["user"], "Skapade offert", f"{quote['id']} — {quote['customer_name']}")
    return jsonify({"ok": True, "quote": quote})


@app.route("/api/quotes/<qid>", methods=["PUT"])
@login_required
def update_quote(qid):
    updates = request.get_json(force=True)
    data = db.load_data("quotes")
    for row in data:
        if str(row.get("id")) == qid:
            if "items" in updates:
                items = updates["items"]
                updates["items"] = json.dumps(items, ensure_ascii=False)
                updates["subtotal"] = round(sum(float(it.get("total", 0)) for it in items), 2)
                updates["moms_total"] = round(sum(
                    float(it.get("total", 0)) * int(it.get("moms", 25)) / (100 + int(it.get("moms", 25)))
                    for it in items
                ), 2)
                updates["total"] = updates["subtotal"]
            row.update(updates)
            row["updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            break
    db.save_data("quotes", data)
    _log_activity(session["user"], "Uppdaterade offert", qid)
    return jsonify({"ok": True})


@app.route("/api/quotes/<qid>", methods=["DELETE"])
@login_required
def delete_quote(qid):
    data = db.load_data("quotes")
    data = [q for q in data if str(q.get("id")) != qid]
    db.save_data("quotes", data)
    _log_activity(session["user"], "Raderade offert", qid)
    return jsonify({"ok": True})


@app.route("/api/quotes/<qid>/status", methods=["PUT"])
@login_required
def update_quote_status(qid):
    d = request.get_json(force=True)
    new_status = d.get("status")
    data = db.load_data("quotes")
    for row in data:
        if str(row.get("id")) == qid:
            row["status"] = new_status
            row["updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            break
    db.save_data("quotes", data)
    _log_activity(session["user"], f"Offert {new_status}", qid)
    return jsonify({"ok": True})


@app.route("/api/quotes/<qid>/pdf")
@login_required
def generate_quote_pdf(qid):
    """Generate a PDF for a quote."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER

    data = db.load_data("quotes")
    quote = None
    for q in data:
        if str(q.get("id")) == qid:
            quote = q
            break
    if not quote:
        return jsonify({"error": "Offert ej hittad"}), 404

    items = quote.get("items", "[]")
    if isinstance(items, str):
        try:
            items = json.loads(items)
        except Exception:
            items = []

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=30*mm, bottomMargin=20*mm,
                            leftMargin=25*mm, rightMargin=25*mm)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='RightAlign', parent=styles['Normal'], alignment=TA_RIGHT))
    styles.add(ParagraphStyle(name='CenterAlign', parent=styles['Normal'], alignment=TA_CENTER, fontSize=10))
    styles.add(ParagraphStyle(name='SmallGray', parent=styles['Normal'], fontSize=8, textColor=colors.gray))

    elements = []

    # Header
    elements.append(Paragraph(f"<b>{quote.get('bolag', 'Unithread')}</b>", styles['Title']))
    elements.append(Spacer(1, 5*mm))

    # Quote title & ID
    elements.append(Paragraph(f"<b>OFFERT</b> {quote['id']}", styles['Heading2']))
    elements.append(Spacer(1, 3*mm))

    # Customer & date info
    info_data = [
        ["Kund:", quote.get("customer_name", "—"),
         "Datum:", quote.get("created", "")[:10]],
        ["Titel:", quote.get("title", ""),
         "Giltig till:", quote.get("valid_until", "—")],
    ]
    info_table = Table(info_data, colWidths=[55, 150, 65, 100])
    info_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.gray),
        ('TEXTCOLOR', (2, 0), (2, -1), colors.gray),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 8*mm))

    if quote.get("description"):
        elements.append(Paragraph(quote["description"], styles['Normal']))
        elements.append(Spacer(1, 5*mm))

    # Items table
    table_header = ['#', 'Beskrivning', 'Antal', 'À-pris', 'Moms %', 'Summa']
    table_data = [table_header]
    for i, item in enumerate(items, 1):
        table_data.append([
            str(i),
            str(item.get("description", "")),
            str(item.get("quantity", 1)),
            f"{float(item.get('unit_price', 0)):,.0f} kr",
            f"{item.get('moms', 25)}%",
            f"{float(item.get('total', 0)):,.0f} kr",
        ])

    col_widths = [25, 180, 45, 70, 45, 70]
    items_table = Table(table_data, colWidths=col_widths)
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 5*mm))

    # Totals
    subtotal = float(quote.get("subtotal", 0))
    moms_total = float(quote.get("moms_total", 0))
    total = float(quote.get("total", 0))
    totals_data = [
        ['', '', '', '', 'Delsumma:', f"{subtotal:,.0f} kr"],
        ['', '', '', '', 'Moms:', f"{moms_total:,.0f} kr"],
        ['', '', '', '', 'TOTALT:', f"{total:,.0f} kr"],
    ]
    totals_table = Table(totals_data, colWidths=col_widths)
    totals_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (-2, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (-2, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (-2, -1), (-1, -1), 11),
        ('LINEABOVE', (-2, -1), (-1, -1), 1, colors.HexColor('#4f46e5')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(totals_table)
    elements.append(Spacer(1, 15*mm))

    elements.append(Paragraph(f"Skapad av {quote.get('created_by', '')} · Unithread", styles['SmallGray']))

    doc.build(elements)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf',
                     download_name=f"Offert_{quote['id']}.pdf", as_attachment=False)


# ---------------------------------------------------------------------------
# Invoices / Fakturor API
# ---------------------------------------------------------------------------

@app.route("/api/invoices", methods=["GET"])
@login_required
def get_invoices():
    data = db.load_data("invoices")
    status = request.args.get("status")
    bolag = request.args.get("bolag")
    if status and status != "Alla":
        data = [inv for inv in data if inv.get("status") == status]
    if bolag and bolag != "Alla":
        data = [inv for inv in data if inv.get("bolag") == bolag]
    data.sort(key=lambda x: x.get("created", ""), reverse=True)
    return jsonify(data)


@app.route("/api/invoices", methods=["POST"])
@login_required
def create_invoice():
    d = request.get_json(force=True)
    items = d.get("items", [])
    subtotal = sum(float(it.get("total", 0)) for it in items)
    moms_total = sum(
        float(it.get("total", 0)) * int(it.get("moms", 25)) / (100 + int(it.get("moms", 25)))
        for it in items
    )

    invoice = {
        "id": f"F-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}",
        "quote_id": d.get("quote_id", ""),
        "customer_id": d.get("customer_id", ""),
        "customer_name": d.get("customer_name", ""),
        "bolag": d.get("bolag", BUSINESSES[0]),
        "title": d.get("title", "Faktura"),
        "description": d.get("description", ""),
        "items": json.dumps(items, ensure_ascii=False),
        "subtotal": round(subtotal, 2),
        "moms_total": round(moms_total, 2),
        "total": round(subtotal, 2),
        "due_date": d.get("due_date", (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")),
        "status": "Obetald",
        "created_by": session["user"],
        "created": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    db.append_row("invoices", invoice)
    _log_activity(session["user"], "Skapade faktura", f"{invoice['id']} — {invoice['customer_name']}")
    return jsonify({"ok": True, "invoice": invoice})


@app.route("/api/invoices/<iid>", methods=["PUT"])
@login_required
def update_invoice(iid):
    updates = request.get_json(force=True)
    data = db.load_data("invoices")
    for row in data:
        if str(row.get("id")) == iid:
            row.update(updates)
            row["updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            break
    db.save_data("invoices", data)
    _log_activity(session["user"], "Uppdaterade faktura", iid)
    return jsonify({"ok": True})


@app.route("/api/invoices/<iid>/status", methods=["PUT"])
@login_required
def update_invoice_status(iid):
    d = request.get_json(force=True)
    new_status = d.get("status")
    data = db.load_data("invoices")
    for row in data:
        if str(row.get("id")) == iid:
            row["status"] = new_status
            row["updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            break
    db.save_data("invoices", data)
    _log_activity(session["user"], f"Faktura {new_status}", iid)
    return jsonify({"ok": True})


@app.route("/api/invoices/<iid>", methods=["DELETE"])
@login_required
def delete_invoice(iid):
    data = db.load_data("invoices")
    data = [inv for inv in data if str(inv.get("id")) != iid]
    db.save_data("invoices", data)
    _log_activity(session["user"], "Raderade faktura", iid)
    return jsonify({"ok": True})


@app.route("/api/quotes/<qid>/convert-to-invoice", methods=["POST"])
@login_required
def convert_quote_to_invoice(qid):
    """Convert an accepted quote to an invoice."""
    quotes = db.load_data("quotes")
    quote = None
    for q in quotes:
        if str(q.get("id")) == qid:
            quote = q
            break
    if not quote:
        return jsonify({"error": "Offert ej hittad"}), 404

    items = quote.get("items", "[]")
    if isinstance(items, str):
        try:
            items = json.loads(items)
        except Exception:
            items = []

    invoice = {
        "id": f"F-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}",
        "quote_id": qid,
        "customer_id": quote.get("customer_id", ""),
        "customer_name": quote.get("customer_name", ""),
        "bolag": quote.get("bolag", BUSINESSES[0]),
        "title": quote.get("title", "Faktura"),
        "description": quote.get("description", ""),
        "items": json.dumps(items, ensure_ascii=False),
        "subtotal": float(quote.get("subtotal", 0)),
        "moms_total": float(quote.get("moms_total", 0)),
        "total": float(quote.get("total", 0)),
        "due_date": (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d"),
        "status": "Obetald",
        "created_by": session["user"],
        "created": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    db.append_row("invoices", invoice)

    # Mark quote as Fakturerad
    quote["status"] = "Fakturerad"
    quote["updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    db.save_data("quotes", quotes)

    _log_activity(session["user"], "Konverterade offert till faktura", f"{qid} → {invoice['id']}")
    return jsonify({"ok": True, "invoice": invoice})


@app.route("/api/invoices/<iid>/pdf")
@login_required
def generate_invoice_pdf(iid):
    """Generate a PDF for an invoice."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER

    data = db.load_data("invoices")
    invoice = None
    for inv in data:
        if str(inv.get("id")) == iid:
            invoice = inv
            break
    if not invoice:
        return jsonify({"error": "Faktura ej hittad"}), 404

    items = invoice.get("items", "[]")
    if isinstance(items, str):
        try:
            items = json.loads(items)
        except Exception:
            items = []

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=30*mm, bottomMargin=20*mm,
                            leftMargin=25*mm, rightMargin=25*mm)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='InvRightAlign', parent=styles['Normal'], alignment=TA_RIGHT))
    styles.add(ParagraphStyle(name='InvCenterAlign', parent=styles['Normal'], alignment=TA_CENTER, fontSize=10))
    styles.add(ParagraphStyle(name='InvSmallGray', parent=styles['Normal'], fontSize=8, textColor=colors.gray))

    elements = []

    # Header
    elements.append(Paragraph(f"<b>{invoice.get('bolag', 'Unithread')}</b>", styles['Title']))
    elements.append(Spacer(1, 5*mm))

    # Invoice title & ID
    elements.append(Paragraph(f"<b>FAKTURA</b> {invoice['id']}", styles['Heading2']))
    elements.append(Spacer(1, 3*mm))

    # Info
    info_data = [
        ["Kund:", invoice.get("customer_name", "—"),
         "Fakturadatum:", invoice.get("created", "")[:10]],
        ["Titel:", invoice.get("title", ""),
         "Förfallodatum:", invoice.get("due_date", "—")],
    ]
    if invoice.get("quote_id"):
        info_data.append(["Offert-ref:", invoice["quote_id"], "Status:", invoice.get("status", "")])

    info_table = Table(info_data, colWidths=[70, 140, 80, 100])
    info_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.gray),
        ('TEXTCOLOR', (2, 0), (2, -1), colors.gray),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 8*mm))

    if invoice.get("description"):
        elements.append(Paragraph(invoice["description"], styles['Normal']))
        elements.append(Spacer(1, 5*mm))

    # Items table
    table_header = ['#', 'Beskrivning', 'Antal', 'À-pris', 'Moms %', 'Summa']
    table_data = [table_header]
    for i, item in enumerate(items, 1):
        table_data.append([
            str(i),
            str(item.get("description", "")),
            str(item.get("quantity", 1)),
            f"{float(item.get('unit_price', 0)):,.0f} kr",
            f"{item.get('moms', 25)}%",
            f"{float(item.get('total', 0)):,.0f} kr",
        ])

    col_widths = [25, 180, 45, 70, 45, 70]
    items_table = Table(table_data, colWidths=col_widths)
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdf4')]),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 5*mm))

    # Totals
    subtotal = float(invoice.get("subtotal", 0))
    moms_total = float(invoice.get("moms_total", 0))
    total = float(invoice.get("total", 0))
    totals_data = [
        ['', '', '', '', 'Delsumma:', f"{subtotal:,.0f} kr"],
        ['', '', '', '', 'Moms:', f"{moms_total:,.0f} kr"],
        ['', '', '', '', 'ATT BETALA:', f"{total:,.0f} kr"],
    ]
    totals_table = Table(totals_data, colWidths=col_widths)
    totals_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (-2, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (-2, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (-2, -1), (-1, -1), 11),
        ('LINEABOVE', (-2, -1), (-1, -1), 1, colors.HexColor('#059669')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(totals_table)
    elements.append(Spacer(1, 10*mm))

    # Payment instructions
    elements.append(Paragraph("<b>Betalningsvillkor</b>", styles['Normal']))
    elements.append(Spacer(1, 2*mm))
    elements.append(Paragraph(f"Förfallodatum: {invoice.get('due_date', '30 dagar')}", styles['Normal']))
    elements.append(Paragraph(f"Fakturanummer: {invoice['id']}", styles['Normal']))
    elements.append(Spacer(1, 10*mm))

    elements.append(Paragraph(f"Skapad av {invoice.get('created_by', '')} · {invoice.get('bolag', 'Unithread')}", styles['InvSmallGray']))

    doc.build(elements)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf',
                     download_name=f"Faktura_{invoice['id']}.pdf", as_attachment=False)


# ---------------------------------------------------------------------------
# Export API — PDF / Excel
# ---------------------------------------------------------------------------

@app.route("/api/export/expenses")
@login_required
def export_expenses():
    fmt = request.args.get("format", "excel")
    bolag = request.args.get("bolag")
    month = request.args.get("month")
    data = db.load_data("expenses")
    if bolag and bolag != "Alla":
        data = [e for e in data if e.get("bolag") == bolag]
    if month:
        data = [e for e in data if str(e.get("datum", "")).startswith(month)]
    data.sort(key=lambda x: x.get("datum", ""), reverse=True)

    if fmt == "excel":
        return _export_excel(data, "Utgifter",
                             ["datum", "bolag", "kategori", "beskrivning", "leverantor", "belopp", "moms_sats", "moms_belopp"],
                             ["Datum", "Bolag", "Kategori", "Beskrivning", "Leverantör", "Belopp", "Moms %", "Momsbelopp"])
    else:
        return _export_pdf_table("Utgiftsrapport", data,
                                 ["datum", "bolag", "kategori", "beskrivning", "belopp"],
                                 ["Datum", "Bolag", "Kategori", "Beskrivning", "Belopp"])


@app.route("/api/export/revenue")
@login_required
def export_revenue():
    fmt = request.args.get("format", "excel")
    bolag = request.args.get("bolag")
    month = request.args.get("month")
    data = db.load_data("revenue")
    if bolag and bolag != "Alla":
        data = [r for r in data if r.get("bolag") == bolag]
    if month:
        data = [r for r in data if str(r.get("datum", "")).startswith(month)]
    data.sort(key=lambda x: x.get("datum", ""), reverse=True)

    if fmt == "excel":
        return _export_excel(data, "Intäkter",
                             ["datum", "bolag", "kategori", "beskrivning", "kund", "belopp"],
                             ["Datum", "Bolag", "Kategori", "Beskrivning", "Kund", "Belopp"])
    else:
        return _export_pdf_table("Intäktsrapport", data,
                                 ["datum", "bolag", "kategori", "beskrivning", "kund", "belopp"],
                                 ["Datum", "Bolag", "Kategori", "Beskrivning", "Kund", "Belopp"])


@app.route("/api/export/customers")
@login_required
def export_customers():
    data = db.load_data("customers")
    return _export_excel(data, "Kunder",
                         ["name", "company", "email", "phone", "bolag", "stage", "value", "assigned_to", "created"],
                         ["Namn", "Företag", "E-post", "Telefon", "Bolag", "Stadie", "Värde", "Ansvarig", "Skapad"])


def _export_excel(data, sheet_title, fields, headers):
    """Generic Excel export."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Header style
    header_fill = PatternFill(start_color="4f46e5", end_color="4f46e5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    border = Border(bottom=Side(style='thin', color='e2e8f0'))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, item in enumerate(data, 2):
        for col_idx, field in enumerate(fields, 1):
            val = item.get(field, "")
            try:
                val = float(val)
            except (ValueError, TypeError):
                pass
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=f"{sheet_title}_{date.today().isoformat()}.xlsx", as_attachment=True)


def _export_pdf_table(title, data, fields, headers):
    """Generic PDF table export."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=20*mm, bottomMargin=15*mm,
                            leftMargin=15*mm, rightMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"<b>{title}</b>", styles['Title']),
        Paragraph(f"Exporterad {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']),
        Spacer(1, 8*mm),
    ]

    table_data = [headers]
    for item in data[:500]:
        row = []
        for field in fields:
            val = item.get(field, "")
            if field == "belopp":
                try:
                    val = f"{float(val):,.0f} kr"
                except (ValueError, TypeError):
                    pass
            row.append(str(val)[:50])
        table_data.append(row)

    n_cols = len(headers)
    col_width = (landscape(A4)[0] - 30*mm) / n_cols
    t = Table(table_data, colWidths=[col_width] * n_cols)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#e2e8f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)

    doc.build(elements)
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf',
                     download_name=f"{title}_{date.today().isoformat()}.pdf", as_attachment=True)


# ---------------------------------------------------------------------------
# Integrations API
# ---------------------------------------------------------------------------

from integrations import get_all_platforms, create_adapter


@app.route("/api/integrations/platforms")
@login_required
def get_integration_platforms():
    """Return metadata about all available integration platforms."""
    return jsonify(get_all_platforms())


@app.route("/api/integrations")
@login_required
def get_integrations():
    """Return all configured integrations (credentials masked)."""
    data = db.load_data("integrations")
    safe = []
    for row in data:
        masked = dict(row)
        # Mask sensitive fields
        for key in ("access_token", "api_key", "client_secret", "refresh_token", "developer_token"):
            if masked.get(key):
                val = str(masked[key])
                masked[key] = val[:4] + "••••" + val[-4:] if len(val) > 8 else "••••••••"
        safe.append(masked)
    return jsonify(safe)


@app.route("/api/integrations", methods=["POST"])
@admin_required
def save_integration():
    """Create or update an integration configuration."""
    d = request.get_json(force=True)
    platform = _sanitize_string(d.get("platform", ""), 30)
    bolag = _sanitize_string(d.get("bolag", BUSINESSES[0]), 50)

    if not platform:
        return jsonify({"ok": False, "error": "Plattform krävs"}), 400
    if bolag not in BUSINESSES:
        return jsonify({"ok": False, "error": "Ogiltigt bolag"}), 400

    # Build config from required fields
    config = {
        "id": f"int_{platform}_{bolag}_{uuid.uuid4().hex[:6]}",
        "platform": platform,
        "bolag": bolag,
        "enabled": True,
        "created": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "last_sync": "",
        "last_sync_status": "",
        "sync_errors": "",
    }
    # Copy credential fields
    adapter_cls = create_adapter(platform, {})
    if not adapter_cls:
        # Get fields from platform metadata
        from integrations import get_adapter_class
        cls = get_adapter_class(platform)
        if cls:
            for field in cls.REQUIRED_FIELDS:
                config[field["key"]] = _sanitize_string(d.get(field["key"], ""), 500)
    else:
        # Shouldn't happen but handle gracefully
        for key, val in d.items():
            if key not in ("platform", "bolag"):
                config[key] = _sanitize_string(str(val), 500)

    # Actually get required fields properly
    from integrations import get_adapter_class
    cls = get_adapter_class(platform)
    if cls:
        for field in cls.REQUIRED_FIELDS:
            config[field["key"]] = d.get(field["key"], "")

    # Check if integration for this platform+bolag already exists → update it
    data = db.load_data("integrations")
    existing = None
    for row in data:
        if row.get("platform") == platform and row.get("bolag") == bolag:
            existing = row
            break

    if existing:
        # Update existing — keep the old ID, merge new creds
        existing.update(config)
        existing["id"] = existing.get("id", config["id"])
        db.save_data("integrations", data)
    else:
        db.append_row("integrations", config)

    _log_activity(session["user"], "Konfigurerade integration", f"{platform} ({bolag})")
    return jsonify({"ok": True, "integration": {**config, "access_token": "***", "api_key": "***"}})


@app.route("/api/integrations/<int_id>", methods=["DELETE"])
@admin_required
def delete_integration(int_id):
    """Remove an integration configuration."""
    data = db.load_data("integrations")
    data = [row for row in data if str(row.get("id")) != int_id]
    db.save_data("integrations", data)
    _log_activity(session["user"], "Raderade integration", int_id)
    return jsonify({"ok": True})


@app.route("/api/integrations/<int_id>/toggle", methods=["PUT"])
@admin_required
def toggle_integration(int_id):
    """Enable or disable an integration."""
    data = db.load_data("integrations")
    for row in data:
        if str(row.get("id")) == int_id:
            current = str(row.get("enabled", "True")).lower() == "true"
            row["enabled"] = not current
            break
    db.save_data("integrations", data)
    return jsonify({"ok": True})


@app.route("/api/integrations/<int_id>/test", methods=["POST"])
@admin_required
def test_integration(int_id):
    """Test connection for an integration."""
    data = db.load_data("integrations")
    config = None
    for row in data:
        if str(row.get("id")) == int_id:
            config = row
            break
    if not config:
        return jsonify({"ok": False, "error": "Integration ej hittad"}), 404

    adapter = create_adapter(config.get("platform", ""), config)
    if not adapter:
        return jsonify({"ok": False, "error": "Okänd plattform"}), 400

    result = adapter.test_connection()
    return jsonify(result)


@app.route("/api/integrations/<int_id>/sync", methods=["POST"])
@admin_required
def sync_integration(int_id):
    """Manually trigger a sync for an integration."""
    d = request.get_json(force=True) if request.is_json else {}
    since_date = d.get("since_date")

    data = db.load_data("integrations")
    config = None
    for row in data:
        if str(row.get("id")) == int_id:
            config = row
            break
    if not config:
        return jsonify({"ok": False, "error": "Integration ej hittad"}), 404

    platform = config.get("platform", "")
    bolag = config.get("bolag", BUSINESSES[0])
    adapter = create_adapter(platform, config)
    if not adapter:
        return jsonify({"ok": False, "error": "Okänd plattform"}), 400

    try:
        result = adapter.sync_data(since_date)
    except Exception as e:
        # Update status
        config["last_sync"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        config["last_sync_status"] = "error"
        config["sync_errors"] = str(e)[:300]
        db.save_data("integrations", data)
        return jsonify({"ok": False, "error": str(e)[:200]}), 500

    added_expenses = 0
    added_revenue = 0
    skipped = 0

    # Load existing to deduplicate by source_id
    existing_expenses = db.load_data("expenses")
    existing_revenue = db.load_data("revenue")
    existing_exp_source_ids = {str(e.get("source_id", "")) for e in existing_expenses if e.get("source_id")}
    existing_rev_source_ids = {str(r.get("source_id", "")) for r in existing_revenue if r.get("source_id")}

    # Add new expenses
    for exp in result.get("expenses", []):
        source_id = exp.get("source_id", "")
        if source_id and source_id in existing_exp_source_ids:
            skipped += 1
            continue
        expense = {
            "id": str(uuid.uuid4())[:8],
            "bolag": bolag,
            "datum": exp.get("datum", date.today().isoformat()),
            "kategori": exp.get("kategori", "Övrigt"),
            "beskrivning": exp.get("beskrivning", ""),
            "leverantor": platform.replace("_", " ").title(),
            "belopp": float(exp.get("belopp", 0)),
            "moms_sats": int(exp.get("moms_sats", 0)),
            "moms_belopp": 0,
            "source": platform,
            "source_id": source_id,
        }
        expense["moms_belopp"] = round(
            expense["belopp"] * expense["moms_sats"] / (100 + expense["moms_sats"]), 2
        ) if expense["moms_sats"] > 0 else 0
        db.append_row("expenses", expense)
        added_expenses += 1

    # Add new revenue
    for rev in result.get("revenue", []):
        source_id = rev.get("source_id", "")
        if source_id and source_id in existing_rev_source_ids:
            skipped += 1
            continue
        revenue_entry = {
            "id": str(uuid.uuid4())[:8],
            "bolag": bolag,
            "datum": rev.get("datum", date.today().isoformat()),
            "kategori": rev.get("kategori", "Övrigt"),
            "beskrivning": rev.get("beskrivning", ""),
            "kund": rev.get("kund", ""),
            "belopp": float(rev.get("belopp", 0)),
            "source": platform,
            "source_id": source_id,
        }
        db.append_row("revenue", revenue_entry)
        added_revenue += 1

    # Update sync status
    config["last_sync"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    config["last_sync_status"] = "success" if not adapter.errors else "partial"
    config["sync_errors"] = "; ".join(adapter.errors) if adapter.errors else ""
    db.save_data("integrations", data)

    _log_activity(
        session["user"], "Synkade integration",
        f"{platform} ({bolag}): +{added_expenses} utgifter, +{added_revenue} intäkter, {skipped} hoppade"
    )
    return jsonify({
        "ok": True,
        "added_expenses": added_expenses,
        "added_revenue": added_revenue,
        "skipped": skipped,
        "errors": adapter.errors,
    })


@app.route("/api/integrations/sync-all", methods=["POST"])
@admin_required
def sync_all_integrations():
    """Trigger sync for all enabled integrations."""
    data = db.load_data("integrations")
    results = []
    for config in data:
        if str(config.get("enabled", "True")).lower() != "true":
            continue
        int_id = config.get("id", "")
        platform = config.get("platform", "")
        bolag = config.get("bolag", "")

        adapter = create_adapter(platform, config)
        if not adapter:
            results.append({"platform": platform, "bolag": bolag, "ok": False, "error": "Okänd plattform"})
            continue

        try:
            last_sync = config.get("last_sync", "")
            since = last_sync[:10] if last_sync else None
            sync_result = adapter.sync_data(since)

            # Deduplicate
            existing_expenses = db.load_data("expenses")
            existing_revenue = db.load_data("revenue")
            exp_ids = {str(e.get("source_id", "")) for e in existing_expenses if e.get("source_id")}
            rev_ids = {str(r.get("source_id", "")) for r in existing_revenue if r.get("source_id")}

            added_exp = 0
            for exp in sync_result.get("expenses", []):
                sid = exp.get("source_id", "")
                if sid and sid in exp_ids:
                    continue
                expense = {
                    "id": str(uuid.uuid4())[:8],
                    "bolag": bolag,
                    "datum": exp.get("datum", date.today().isoformat()),
                    "kategori": exp.get("kategori", "Övrigt"),
                    "beskrivning": exp.get("beskrivning", ""),
                    "leverantor": platform.replace("_", " ").title(),
                    "belopp": float(exp.get("belopp", 0)),
                    "moms_sats": int(exp.get("moms_sats", 0)),
                    "moms_belopp": 0,
                    "source": platform,
                    "source_id": sid,
                }
                db.append_row("expenses", expense)
                added_exp += 1

            added_rev = 0
            for rev in sync_result.get("revenue", []):
                sid = rev.get("source_id", "")
                if sid and sid in rev_ids:
                    continue
                revenue_entry = {
                    "id": str(uuid.uuid4())[:8],
                    "bolag": bolag,
                    "datum": rev.get("datum", date.today().isoformat()),
                    "kategori": rev.get("kategori", "Övrigt"),
                    "beskrivning": rev.get("beskrivning", ""),
                    "kund": rev.get("kund", ""),
                    "belopp": float(rev.get("belopp", 0)),
                    "source": platform,
                    "source_id": sid,
                }
                db.append_row("revenue", revenue_entry)
                added_rev += 1

            config["last_sync"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            config["last_sync_status"] = "success"
            config["sync_errors"] = ""
            results.append({"platform": platform, "bolag": bolag, "ok": True,
                            "added_expenses": added_exp, "added_revenue": added_rev})

        except Exception as e:
            config["last_sync"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            config["last_sync_status"] = "error"
            config["sync_errors"] = str(e)[:300]
            results.append({"platform": platform, "bolag": bolag, "ok": False, "error": str(e)[:200]})

    db.save_data("integrations", data)
    return jsonify({"ok": True, "results": results})


@app.route("/api/integrations/summary")
@login_required
def get_integrations_summary():
    """Return a summary of integration data for dashboard display."""
    integrations = db.load_data("integrations")
    expenses = db.load_data("expenses")
    revenue = db.load_data("revenue")

    current_month = datetime.now().strftime("%Y-%m")
    platforms = {}

    for integ in integrations:
        platform = integ.get("platform", "")
        bolag = integ.get("bolag", "")
        enabled = str(integ.get("enabled", "True")).lower() == "true"

        if platform not in platforms:
            from integrations import get_adapter_class
            cls = get_adapter_class(platform)
            platforms[platform] = {
                "platform": platform,
                "display_name": cls.DISPLAY_NAME if cls else platform,
                "icon": cls.ICON if cls else "🔗",
                "enabled": enabled,
                "last_sync": integ.get("last_sync", ""),
                "last_sync_status": integ.get("last_sync_status", ""),
                "businesses": [],
                "month_expenses": 0,
                "month_revenue": 0,
                "total_expenses": 0,
                "total_revenue": 0,
            }

        platforms[platform]["businesses"].append(bolag)

        # Calculate totals from expenses/revenue with matching source
        p_exp = [e for e in expenses if e.get("source") == platform and e.get("bolag") == bolag]
        p_rev = [r for r in revenue if r.get("source") == platform and r.get("bolag") == bolag]

        platforms[platform]["total_expenses"] += sum(float(e.get("belopp", 0)) for e in p_exp)
        platforms[platform]["total_revenue"] += sum(float(r.get("belopp", 0)) for r in p_rev)
        platforms[platform]["month_expenses"] += sum(
            float(e.get("belopp", 0)) for e in p_exp if str(e.get("datum", "")).startswith(current_month)
        )
        platforms[platform]["month_revenue"] += sum(
            float(r.get("belopp", 0)) for r in p_rev if str(r.get("datum", "")).startswith(current_month)
        )

    return jsonify(list(platforms.values()))


# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    app.config["UPLOAD_FOLDER"].mkdir(parents=True, exist_ok=True)
    app.config["PROJECT_UPLOAD_FOLDER"].mkdir(parents=True, exist_ok=True)
    socketio.run(app, debug=True, port=5000)
